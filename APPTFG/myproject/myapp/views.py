import openpyxl

from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from .models import Objetivo, Indicador, AGC, UGC, AcuerdoIndicadores
from .forms import ObjetivoForm, IndicadorForm, AGCForm, UGCForm, AcuerdoIndicadoresForm, UploadExcelForm
from openpyxl.utils import get_column_letter



def inicio(request):
    return render(request, 'myapp/inicio.html')


# Gestionar UGC

class GestionarUGCView(View):
    def get(self, request):
        ugcs = UGC.objects.all()
        return render(request, 'myapp/ugc.html', {'ugcs': ugcs})

    def post(self, request):
        if 'delete' in request.POST:
            id_ugc = request.POST.get('id_ugc')
            ugc = get_object_or_404(UGC, id_ugc=id_ugc)
            ugc.delete()
            messages.success(request, 'UGC eliminado con éxito.')
        return redirect('gestionar_ugc')

class CrearUGCView(View):
    def get(self, request):
        form = UGCForm()
        return render(request, 'myapp/crear_ugc.html', {'form': form})

    def post(self, request):
        form = UGCForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'UGC creado con éxito.')
            return redirect('gestionar_ugc')
        else:
            messages.error(request, 'Error al crear UGC.')
            return render(request, 'myapp/crear_ugc.html', {'form': form})

class EditarUGCView(View):
    def get(self, request, id_ugc):
        ugc = get_object_or_404(UGC, id_ugc=id_ugc)
        form = UGCForm(instance=ugc)
        return render(request, 'myapp/editar_ugc.html', {'form': form, 'id_ugc': id_ugc})

    def post(self, request, id_ugc):
        ugc = get_object_or_404(UGC, id_ugc=id_ugc)
        form = UGCForm(request.POST, instance=ugc)
        if form.is_valid():
            form.save()
            messages.success(request, 'UGC editado con éxito.')
            return redirect('gestionar_ugc')
        else:
            messages.error(request, 'Error al editar UGC.')
            return render(request, 'myapp/editar_ugc.html', {'form': form, 'id_ugc': id_ugc})



# Gestionar objetivos
class GestionarObjetivosView(View):
    def get(self, request):
        objetivos = Objetivo.objects.all()
        return render(request, 'myapp/objetivos.html', {'objetivos': objetivos})

    def post(self, request):
        if 'delete' in request.POST:
            id_objetivo = request.POST.get('id_objetivo')
            objetivo = get_object_or_404(Objetivo, id_objetivo=id_objetivo)
            objetivo.delete()
            messages.success(request, 'Objetivo eliminado con éxito.')
        return redirect('gestionar_objetivos')

class CrearObjetivoView(View):
    def get(self, request):
        form = ObjetivoForm()
        return render(request, 'myapp/crear_objetivo.html', {'form': form})

    def post(self, request):
        form = ObjetivoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Objetivo creado con éxito.')
            return redirect('gestionar_objetivos')
        else:
            messages.error(request, 'Error al crear Objetivo.')
            return render(request, 'myapp/crear_objetivo.html', {'form': form})

class EditarObjetivoView(View):
    def get(self, request, id_objetivo):
        objetivo = get_object_or_404(Objetivo, id_objetivo=id_objetivo)
        form = ObjetivoForm(instance=objetivo)
        return render(request, 'myapp/editar_objetivo.html', {'form': form, 'id_objetivo': id_objetivo})

    def post(self, request, id_objetivo):
        objetivo = get_object_or_404(Objetivo, id_objetivo=id_objetivo)
        form = ObjetivoForm(request.POST, instance=objetivo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Objetivo editado con éxito.')
            return redirect('gestionar_objetivos')
        else:
            messages.error(request, 'Error al editar Objetivo.')
            return render(request, 'myapp/editar_objetivo.html', {'form': form, 'id_objetivo': id_objetivo})
        

# Gestionar indicadores

class GestionarIndicadoresView(View):
    def get(self, request):
        indicadores = Indicador.objects.all()
        return render(request, 'myapp/indicadores.html', {'indicadores': indicadores})

class CrearIndicadorView(View):
    def get(self, request):
        form = IndicadorForm()
        return render(request, 'myapp/crear_indicador.html', {'form': form})

    def post(self, request):
        form = IndicadorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Indicador creado con éxito.')
            return redirect('gestionar_indicadores')
        else:
            messages.error(request, 'Error al crear indicador.')
            return render(request, 'myapp/crear_indicador.html', {'form': form})

class EditarIndicadorView(View):
    def get(self, request, pk):
        indicador = get_object_or_404(Indicador, pk=pk)
        form = IndicadorForm(instance=indicador)
        return render(request, 'myapp/editar_indicador.html', {'form': form, 'indicador': indicador})

    def post(self, request, pk):
        indicador = get_object_or_404(Indicador, pk=pk)
        form = IndicadorForm(request.POST, instance=indicador)
        if form.is_valid():
            form.save()
            messages.success(request, 'Indicador editado con éxito.')
            return redirect('gestionar_indicadores')
        else:
            messages.error(request, 'Error al editar indicador.')
            return render(request, 'myapp/editar_indicador.html', {'form': form, 'indicador': indicador})

class EliminarIndicadorView(View):
    def post(self, request, pk):
        indicador = get_object_or_404(Indicador, pk=pk)
        indicador.delete()
        messages.success(request, 'Indicador eliminado con éxito.')
        return redirect('gestionar_indicadores')
    

# Gestionar AGC

class GestionarAGCView(View):
    def get(self, request):
        agcs = AGC.objects.all()
        return render(request, 'myapp/agc.html', {'agcs': agcs})

class CrearAGCView(View):
    def get(self, request):
        form = AGCForm()
        return render(request, 'myapp/crear_agc.html', {'form': form})

    def post(self, request):
        form = AGCForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'AGC creado con éxito.')
            return redirect('gestionar_agc')
        else:
            messages.error(request, 'Error al crear AGC.')
            return render(request, 'myapp/crear_agc.html', {'form': form})

class EditarAGCView(View):
    def get(self, request, pk):
        agc = get_object_or_404(AGC, pk=pk)
        form = AGCForm(instance=agc)
        indicadores = AcuerdoIndicadores.objects.filter(id_agc=agc)
        return render(request, 'myapp/editar_agc.html', {'form': form, 'agc': agc, 'indicadores': indicadores})

    def post(self, request, pk):
        agc = get_object_or_404(AGC, pk=pk)
        form = AGCForm(request.POST, instance=agc)
        if form.is_valid():
            form.save()
            messages.success(request, 'AGC editado con éxito.')
            return redirect('gestionar_agc')
        else:
            messages.error(request, 'Error al editar AGC.')
            indicadores = AcuerdoIndicadores.objects.filter(id_agc=agc)
            return render(request, 'myapp/editar_agc.html', {'form': form, 'agc': agc, 'indicadores': indicadores})

class EliminarAGCView(View):
    def post(self, request, pk):
        agc = get_object_or_404(AGC, pk=pk)
        agc.delete()
        messages.success(request, 'AGC eliminado con éxito.')
        return redirect('gestionar_agc')

class GestionarIndicadoresAGCView(View):
    def get(self, request, pk):
        agc = get_object_or_404(AGC, pk=pk)
        indicadores = AcuerdoIndicadores.objects.filter(id_agc=agc)
        return render(request, 'myapp/editar_agc.html', {'agc': agc, 'indicadores': indicadores})

class CrearAcuerdoIndicadorView(View):
    def get(self, request, agc_pk):
        form = AcuerdoIndicadoresForm(initial={'id_agc': agc_pk})
        return render(request, 'myapp/crear_acuerdo_indicador.html', {'form': form, 'agc_pk': agc_pk})

    def post(self, request, agc_pk):
        form = AcuerdoIndicadoresForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Acuerdo de Indicador creado con éxito.')
            return redirect('gestionar_indicadores_agc', pk=agc_pk)
        else:
            messages.error(request, 'Error al crear Acuerdo de Indicador.')
            return render(request, 'myapp/crear_acuerdo_indicador.html', {'form': form, 'agc_pk': agc_pk})

class EditarAcuerdoIndicadorView(View):
    def get(self, request, pk):
        acuerdo_indicador = get_object_or_404(AcuerdoIndicadores, pk=pk)
        form = AcuerdoIndicadoresForm(instance=acuerdo_indicador)
        return render(request, 'myapp/editar_acuerdo_indicador.html', {'form': form, 'acuerdo_indicador': acuerdo_indicador})

    def post(self, request, pk):
        acuerdo_indicador = get_object_or_404(AcuerdoIndicadores, pk=pk)
        form = AcuerdoIndicadoresForm(request.POST, instance=acuerdo_indicador)
        if form.is_valid():
            form.save()
            messages.success(request, 'Acuerdo de Indicador editado con éxito.')
            return redirect('gestionar_indicadores_agc', pk=acuerdo_indicador.id_agc.pk)
        else:
            messages.error(request, 'Error al editar Acuerdo de Indicador.')
            return render(request, 'myapp/editar_acuerdo_indicador.html', {'form': form, 'acuerdo_indicador': acuerdo_indicador})

class EliminarAcuerdoIndicadorView(View):
    def post(self, request, pk):
        acuerdo_indicador = get_object_or_404(AcuerdoIndicadores, pk=pk)
        agc_pk = acuerdo_indicador.id_agc.pk
        acuerdo_indicador.delete()
        messages.success(request, 'Acuerdo de Indicador eliminado con éxito.')
        return redirect('gestionar_indicadores_agc', pk=agc_pk)
    

from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy

class CustomLoginView(LoginView):
    template_name = 'myapp/login.html'
    success_url = reverse_lazy('inicio')




def export_agc_excel(request):
    # Crear un libro de trabajo y una hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AGC e Indicadores'

    # Definir los encabezados de las columnas
    headers = [
        'Año AGC', 'UGC', 'Descripción AGC', 'Nombre Indicador', 
        'Descripción Indicador', 'Fuente', 'Nombre Objetivo', 
        'Descripción Objetivo', 'Perspectiva', 'Valor Mínimo', 
        'Valor Óptimo', 'Valor Obtenido', 'Peso', 'Resultado', '% Conseguido'
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Obtener todos los AGC y sus indicadores asociados
    row_num = 2
    for agc in AGC.objects.all():
        acuerdos = AcuerdoIndicadores.objects.filter(id_agc=agc)
        if not acuerdos:
            ws[f'A{row_num}'] = agc.anio_agc
            ws[f'B{row_num}'] = agc.id_ugc.nombre_ugc
            ws[f'C{row_num}'] = agc.descripcion_agc
            row_num += 1
        else:
            for acuerdo in acuerdos:
                indicador = acuerdo.id_indicador
                objetivo = indicador.id_objetivo
                
                ws[f'A{row_num}'] = agc.anio_agc
                ws[f'B{row_num}'] = agc.id_ugc.nombre_ugc
                ws[f'C{row_num}'] = agc.descripcion_agc
                ws[f'D{row_num}'] = indicador.nombre_indicador
                ws[f'E{row_num}'] = indicador.descripcion_indicador
                ws[f'F{row_num}'] = indicador.fuente
                ws[f'G{row_num}'] = objetivo.nombre_objetivo
                ws[f'H{row_num}'] = objetivo.descripcion_objetivo
                ws[f'I{row_num}'] = objetivo.perspectiva
                ws[f'J{row_num}'] = acuerdo.valor_min
                ws[f'K{row_num}'] = acuerdo.valor_opt
                ws[f'L{row_num}'] = acuerdo.valor_obtenido
                ws[f'M{row_num}'] = acuerdo.peso_indicador
                ws[f'N{row_num}'] = acuerdo.resultado_indicador
                ws[f'O{row_num}'] = acuerdo.porcentaje_conseguido
                row_num += 1

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=agc_indicadores.xlsx'
    return response

def import_agc_excel(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                agc_anio = row[0]
                indicador_nombre = row[3]
                valor_obtenido = row[11]

                acuerdos = AcuerdoIndicadores.objects.filter(
                    id_agc__anio_agc=agc_anio,
                    id_indicador__nombre_indicador=indicador_nombre
                )

                if acuerdos.exists():
                    for acuerdo in acuerdos:
                        acuerdo.valor_obtenido = valor_obtenido
                        acuerdo.save()

            return redirect('gestionar_agc')
    else:
        form = UploadExcelForm()

    return render(request, 'myapp/import_agc_excel.html', {'form': form})


def import_export(request):
    return render(request, 'myapp/import_export.html')

def export_objetivo_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Agrega los encabezados de las columnas
    ws.append(['Nombre Objetivo', 'Descripción Objetivo', 'Perspectiva'])

    # Configura la respuesta HTTP para la descarga del archivo
    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_objetivos.xlsx'
    return response

def import_objetivo_excel(request):
    if request.method == 'POST':
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre_objetivo = row[0]
                descripcion_objetivo = row[1]
                perspectiva = row[2]

                # Crea o actualiza el registro de Objetivo
                objetivo, created = Objetivo.objects.update_or_create(
                    nombre_objetivo=nombre_objetivo,
                    defaults={
                        'descripcion_objetivo': descripcion_objetivo,
                        'perspectiva': perspectiva,
                    }
                )

            return redirect('gestionar_objetivos')
    else:
        form = UploadExcelForm()

    return render(request, 'myapp/import_objetivo_excel.html', {'form': form})